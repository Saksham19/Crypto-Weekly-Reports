import requests as req
import datetime as dt
import pandas as pd
import os
import datetime 
from datetime import date
import re
import requests
import urllib.request, json 


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,index=True,header=True,**to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
    	# try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
# write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs,index=index,header=header)

    # save the workbook
    writer.save()


def date_to_unix(
    year = 2010,
    month = 1,
    day = 15,
    ):
    '''Returns the date (UTC time) in Unix time'''

    time = int((dt.datetime(year, month, day, 0, 0, 0).timestamp()))
    return time


def glassnode(
    endpoint, 
    start,
    until, 
    api_key = '1e736fd8-effa-413a-8732-c9f0c2df458a', # Insert your own API key here, this one is only illustrative.
    asset = 'BTC',
    status = False,
    headers = False,
    resolution = '24h',
    wait = 10
    ):
    '''Returns a Dataframe of time, value pairs for a metric from the Glassnode API.
    
    Parameters
    ----------
    endpoint : str
        Endpoint url after https://api.glassnode.com, corresponding to some metric (ex. '/v1/metrics/indicators/puell_multiple' )
    start : list
        Start date as a list in decreasing order (ex. [2015, 11, 27] )
    until : list
        End date as a list in decreasing order (ex. [2018, 5, 13] )
    api_key : str
        Your API key (ex. 'a2b123be-2c50-4dc9-bdf4-cded52c3d1fc' )
    asset : str
        Asset to which the metric refers. (ex. BTC )
    status : bool
        Option to print HTTP status code. '<Response [200]>' means success.
    headers : bool
        Option to print HTTP headers. Contains REST API request and response metadata.
    resolution : str
        Temporal resolution of the data received. Can be '10m', '1h', '24h', '1w' or '1month'.
    wait : float
        Seconds until the connection timeouts. ALWAYS specify a period.
    Returns
    -------
    DataFrame
        List of {'t' : unix-time, 'v' : 'metric-value'} pairs
    '''

    s = date_to_unix(year=start[0], month=start[1], day=start[2])
    u = date_to_unix(year=until[0], month=until[1], day=until[2])

    response = req.get(
        f'https://api.glassnode.com{endpoint}', 
        {
        'api_key': api_key, 
        'a': asset, 
        's': s, 
        'u': u,
        'i': resolution
        }, 
        timeout = wait)

    df = pd.DataFrame(response.json())

    if status:
        print(response)
    if headers:
        print(response.headers)
    return df

cwd = os.getcwd()
date_month=date.today()
date_today=date.today()-datetime.timedelta(50)

current_year=date_month.year
current_month=(date_month.month-1)
print(current_month)
print([date_month.year,date_month.month,date_month.day])








##########SOPR##############################3


sopr=glassnode("/v1/metrics/indicators/sopr",[date_month.year,date_month.month-11,date_month.day],[date_month.year,date_month.month,date_month.day])

sopr_df=pd.DataFrame(sopr)
sopr_df.columns=["DATE","SOPR"]
sopr_df["DATE"]=pd.to_datetime(sopr_df['DATE'], unit='s')
print(sopr_df)

sopr_df.to_csv(cwd+"/IGNORE/SOPR_FINAL.csv",index=False,header=True,index_label="DATE")


x3 = pd.read_csv(cwd+"/IGNORE/BASE/SOPR_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/SOPR_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


sopr_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(sopr_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_sr2= df4[~df4.index.duplicated()]

df_sr2.to_csv(cwd+"/IGNORE/SOPR_FINAL.csv",index=True,index_label="DATE")
df_sr2.to_csv(cwd+"/IGNORE/BASE/SOPR_BASE.csv",index=True,index_label="DATE")





##############Stock to Flow model############################
###GLASSNODE
stf=glassnode("/v1/metrics/indicators/stock_to_flow_ratio",[date_month.year,date_month.month-11,date_month.day],[date_month.year,date_month.month,date_month.day])

stf_df=pd.DataFrame(stf)
stf_df.columns=["DATE","STF"]
stf_df["DATE"]=pd.to_datetime(stf_df['DATE'], unit='s')
#print(stf_df["STF"])

stf_df=pd.concat([stf_df.drop(['STF'], axis=1), stf_df['STF'].apply(pd.Series)], axis=1)
del(stf_df['daysTillHalving'])
stf_df.columns=["DATE","SF_RATIO"]
print(stf_df)

stf_df.to_csv(cwd+"/IGNORE/STF_FINAL.csv",index=False,header=True,index_label="DATE")


x3 = pd.read_csv(cwd+"/IGNORE/BASE/STF_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/STF_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


stf_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(stf_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_stf2= df4[~df4.index.duplicated()]

df_stf2.to_csv(cwd+"/IGNORE/STF_FINAL.csv",index=True,index_label="DATE")
df_stf2.to_csv(cwd+"/IGNORE/BASE/STF_BASE.csv",index=True,index_label="DATE")



##################NETWORK_VALUE_TO_TRANSACTIONS SIGNALS##################################
#blockchain.com


with urllib.request.urlopen("https://api.blockchain.info/charts/nvts?timespan=1year&sampled=true&metadata=false&cors=true&format=json") as url:
    data = json.loads(url.read().decode())

#print(data["values"])

nvts=[]

for	i in range(0,len(data["values"])):
	time=data["values"][i]["x"]
	ts=int(time)
	time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
	value=data["values"][i]['y']
	nvts.append([time,value])

#print(nvts)

df=pd.DataFrame(nvts,columns=['DATE','NVTS'])
df=df.set_index("DATE")
df= df[~df.index.duplicated(keep="last")]

df.to_csv(cwd+"/IGNORE/NVTS_FINAL.csv",index=True,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/IGNORE/BASE/NVTS_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/NVTS_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


nvts_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(nvts_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_nv2= df4[~df4.index.duplicated()]

df_nv2.to_csv(cwd+"/IGNORE/NVTS_FINAL.csv",index=True,index_label="DATE")
df_nv2.to_csv(cwd+"/IGNORE/BASE/NVTS_BASE.csv",index=True,index_label="DATE")

##################Market Value to Realized Value ##################################
#blockchain.com


with urllib.request.urlopen("https://api.blockchain.info/charts/mvrv?timespan=1year&sampled=true&metadata=false&cors=true&format=json") as url:
    data = json.loads(url.read().decode())

#print(data["values"])

mvrv=[]

for	i in range(0,len(data["values"])):
	time=data["values"][i]["x"]
	ts=int(time)
	time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
	value=data["values"][i]['y']
	mvrv.append([time,value])

#print(mvrv)

df=pd.DataFrame(mvrv,columns=['DATE','MVRV'])
df=df.set_index("DATE")
df= df[~df.index.duplicated(keep="last")]

df.to_csv(cwd+"/IGNORE/MVRV_FINAL.csv",index=True,header=True,index_label="DATE")


x3 = pd.read_csv(cwd+"/IGNORE/BASE/MVRV_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/MVRV_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


mvrv_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(mvrv_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_mv2= df4[~df4.index.duplicated()]

df_mv2.to_csv(cwd+"/IGNORE/MVRV_FINAL.csv",index=True,index_label="DATE")
df_mv2.to_csv(cwd+"/IGNORE/BASE/MVRV_BASE.csv",index=True,index_label="DATE")


####Exchange Traded Volume###############################
#blockchain.com


with urllib.request.urlopen("https://api.blockchain.info/charts/trade-volume?timespan=1year&sampled=true&metadata=false&cors=true&format=json") as url:
    data = json.loads(url.read().decode())

#print(data["values"])

etv=[]

for	i in range(0,len(data["values"])):
	time=data["values"][i]["x"]
	ts=int(time)
	time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
	value=data["values"][i]['y']
	etv.append([time,value])

#print(mvrv)

df=pd.DataFrame(etv,columns=['DATE','ETV'])
df=df.set_index("DATE")
df= df[~df.index.duplicated(keep="last")]

df.to_csv(cwd+"/IGNORE/ETV_FINAL.csv",index=True,header=True,index_label="DATE")


x3 = pd.read_csv(cwd+"/IGNORE/BASE/ETV_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/ETV_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


etv_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(etv_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_ev2= df4[~df4.index.duplicated()]

df_ev2.to_csv(cwd+"/IGNORE/ETV_FINAL.csv",index=True,index_label="DATE")
df_ev2.to_csv(cwd+"/IGNORE/BASE/ETV_BASE.csv",index=True,index_label="DATE")


##############BTC ACTIVE ADDRESSES######################################
#GLASSNODE



btc_aa=glassnode("/v1/metrics/addresses/active_count",[date_month.year,date_month.month-11,date_month.day],[date_month.year,date_month.month,date_month.day])

btc_aa_df=pd.DataFrame(btc_aa)
btc_aa_df.columns=["DATE","Active Address"]
btc_aa_df["DATE"]=pd.to_datetime(btc_aa_df['DATE'], unit='s')



btc_aa_df.to_csv(cwd+"/IGNORE/BTC_AA_FINAL.csv",index=False,header=True,index_label="DATE")


x3 = pd.read_csv(cwd+"/IGNORE/BASE/BTC_AA_BASE.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/IGNORE/BTC_AA_FINAL.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


btc_aa_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(btc_aa_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_aa2= df4[~df4.index.duplicated()]

df_aa2.to_csv(cwd+"/IGNORE/BTC_AA_FINAL.csv",index=True,index_label="DATE")
df_aa2.to_csv(cwd+"/IGNORE/BASE/BTC_AA_BASE.csv",index=True,index_label="DATE")











####excel_work################################

new=pd.concat([df_sr2,df_stf2,df_nv2,df_mv2,df_ev2,df_aa2],axis=1)

append_df_to_excel(cwd+"//OI_COMBINED//ON_CHAIN_COMBINED.xlsx",new,sheet_name="Sheet1",startrow=0)
