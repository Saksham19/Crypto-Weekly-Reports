import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from time import sleep
import datetime as dt
import pandas as pd
import datetime
from datetime import date
import openpyxl
import xlrd
from xlrd import open_workbook
from heapq import merge 
import os
import requests
import urllib.request, json 
import yfinance as yf
import pandas_datareader as pdr
from openpyxl import load_workbook
import xlsxwriter 


cwd = os.getcwd()
date_month=date.today()
date_today=date.today()-datetime.timedelta(50)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,index=True,header=True,**to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
    	# try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
# write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs,index=index,header=header)

    # save the workbook
    writer.save()



###list of exchanges (based on bybit - open interest available)
#Done manually as should stay the same
#https://www.bybt.com/BitcoinOpenInterest

exchange_list=["Binance","Bybit","Okex","Huobi"] #huobipro #
exchange_list = ','.join(exchange_list)






#######################VOLUME DATA - CRYPTOCOMPARE###############################################################

#Binance data

with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=BINANCE") as url:
    data = json.loads(url.read().decode())
    



binance_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    binance_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(binance_vol,columns=['DATE','Binance Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Binance_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Binance_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Binance_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


binance_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(binance_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_bn2= df4[~df4.index.duplicated()]

df_bn2.to_csv(cwd+"/Volume_Data/IGNORE/Binance_Volume_Final.csv",index=True,index_label="DATE")
df_bn2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Binance_Volume_Base.csv",index=True,index_label="DATE")



###okex



with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=okex") as url:
    data = json.loads(url.read().decode())
    



okex_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    okex_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(okex_vol,columns=['DATE','okex Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/okex_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/okex_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/okex_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


okex_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(okex_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_ok2= df4[~df4.index.duplicated()]

df_ok2.to_csv(cwd+"/Volume_Data/IGNORE/okex_Volume_Final.csv",index=True,index_label="DATE")
df_ok2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/okex_Volume_Base.csv",index=True,index_label="DATE")



####Coinbase



with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=coinbase") as url:
    data = json.loads(url.read().decode())
    



coinbase_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    coinbase_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(coinbase_vol,columns=['DATE','Coinbase Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Coinbase_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Coinbase_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Coinbase_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


coinbase_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(coinbase_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_cb2= df4[~df4.index.duplicated()]

df_cb2.to_csv(cwd+"/Volume_Data/IGNORE/Coinbase_Volume_Final.csv",index=True,index_label="DATE")
df_cb2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Coinbase_Volume_Base.csv",index=True,index_label="DATE")



###bitfinex

with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=bitfinex") as url:
    data = json.loads(url.read().decode())
    


bitfinex_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    bitfinex_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(bitfinex_vol,columns=['DATE','Bitfinex Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Bitfinex_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Bitfinex_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Bitfinex_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


bitfinex_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(bitfinex_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_bf2= df4[~df4.index.duplicated()]

df_bf2.to_csv(cwd+"/Volume_Data/IGNORE/Bitfinex_Volume_Final.csv",index=True,index_label="DATE")
df_bf2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Bitfinex_Volume_Base.csv",index=True,index_label="DATE")




###huobi

with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=huobipro") as url:
    data = json.loads(url.read().decode())
    


huobi_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    huobi_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(huobi_vol,columns=['DATE','Huobi Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Huobi_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Huobi_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Huobi_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


huobi_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(huobi_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_hb2= df4[~df4.index.duplicated()]

df_hb2.to_csv(cwd+"/Volume_Data/IGNORE/Huobi_Volume_Final.csv",index=True,index_label="DATE")
df_hb2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Huobi_Volume_Base.csv",index=True,index_label="DATE")




##Liquid


with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=Liquid") as url:
    data = json.loads(url.read().decode())
    


liquid_vol=[]
for i in range(0,49):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    liquid_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(liquid_vol,columns=['DATE','Liquid Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Liquid_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Liquid_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Liquid_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


liquid_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(liquid_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_lq2= df4[~df4.index.duplicated()]

df_lq2.to_csv(cwd+"/Volume_Data/IGNORE/Liquid_Volume_Final.csv",index=True,index_label="DATE")
df_lq2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Liquid_Volume_Base.csv",index=True,index_label="DATE")




##Gemini


with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=Gemini") as url:
    data = json.loads(url.read().decode())
    


gemini_vol=[]
for i in range(0,len(data["Data"])):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    gemini_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(gemini_vol,columns=['DATE','Gemini Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Gemini_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Gemini_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Gemini_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


gemini_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(gemini_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_gm2= df4[~df4.index.duplicated()]

df_gm2.to_csv(cwd+"/Volume_Data/IGNORE/Gemini_Volume_Final.csv",index=True,index_label="DATE")
df_gm2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Gemini_Volume_Base.csv",index=True,index_label="DATE")



##Kraken


with urllib.request.urlopen("https://min-api.cryptocompare.com/data/exchange/symbol/histoday?fsym=BTC&tsym=USD&limit=365&e=Kraken") as url:
    data = json.loads(url.read().decode())
    


kraken_vol=[]
for i in range(0,len(data["Data"])):
    time=data["Data"][i]['time']
    ts = int(time)
    time=dt.datetime.utcfromtimestamp(ts).strftime('%d-%m-%Y')
    value=data["Data"][i]['volumetotal']
    kraken_vol.append([time,value])
#print(data_mvda)
df=pd.DataFrame(kraken_vol,columns=['DATE','Kraken Volume'])

df.to_csv(cwd+"/Volume_Data/IGNORE/Kraken_Volume_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Kraken_Volume_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Kraken_Volume_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


kraken_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(kraken_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_kk2= df4[~df4.index.duplicated()]

df_kk2.to_csv(cwd+"/Volume_Data/IGNORE/Kraken_Volume_Final.csv",index=True,index_label="DATE")
df_kk2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Kraken_Volume_Base.csv",index=True,index_label="DATE")


new = pd.concat([df_bn2,df_ok2,df_cb2,df_bf2,df_hb2,df_lq2,df_gm2,df_kk2], axis=1) 

print(new)

append_df_to_excel(cwd+"//Volume_Data//Volumes_Combined//Volumes_Open_Interest_Combined.xlsx",new,sheet_name="VOLUME_DAILY",startrow=0)



##################################################OPEN_INTEREST_DATA###########################################################################################



with urllib.request.urlopen("https://fapi.bybt.com/api/openInterest/v3/chart?symbol=BTC&timeType=0&exchangeName=&type=0") as url:
    data = json.loads(url.read().decode())


##GETTING THE DATES
date_base=(data["data"]["dateList"])
time=data["data"]["dateList"]

time_new=[]
for i in range(0,len(time)):

	time_actual=dt.datetime.utcfromtimestamp(time[i]/1000).strftime('%d-%m-%Y')
	time_new.append(time_actual)



df_time=pd.DataFrame(time_new,columns=['DATE'])





############BINANCE_OI###############################


value=data["data"]["dataMap"]["Binance"]

#binance_oi=[time,value]
#binance_oi.append([time_new,value])

df_value=pd.DataFrame(value,columns=["BINANCE OI"])

binance_df=pd.concat([df_time,df_value],axis=1)

binance_df.to_csv(cwd+"/Volume_Data/IGNORE/Binance_OI_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/Binance_OI_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/Binance_OI_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


binance_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(binance_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_bno2= df4[~df4.index.duplicated()]

df_bno2.to_csv(cwd+"/Volume_Data/IGNORE/Binance_OI_Final.csv",index=True,index_label="DATE")
df_bno2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/Binance_OI_Base.csv",index=True,index_label="DATE")



#############CME_OI#########################################


value=data["data"]["dataMap"]["CME"]

#cme_oi=[]
#cme_oi.append([time_new,value])

df_value=pd.DataFrame(value,columns=["CME OI"])

cme_df=pd.concat([df_time,df_value],axis=1)

cme_df.to_csv(cwd+"/Volume_Data/IGNORE/CME_OI_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/CME_OI_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/CME_OI_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


cme_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(cme_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_cmeo2= df4[~df4.index.duplicated()]

df_cmeo2.to_csv(cwd+"/Volume_Data/IGNORE/CME_OI_Final.csv",index=True,index_label="DATE")
df_cmeo2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/CME_OI_Base.csv",index=True,index_label="DATE")






##################OKEX_OI#########################################


value=data["data"]["dataMap"]["Okex"]

#okex_oi=[]
#okex_oi.append([time_new,value])


df_value=pd.DataFrame(value,columns=["OKEX OI"])

okex_df=pd.concat([df_time,df_value],axis=1)

okex_df.to_csv(cwd+"/Volume_Data/IGNORE/OKEX_OI_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/OKEX_OI_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/OKEX_OI_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


okex_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(okex_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_okexo2= df4[~df4.index.duplicated()]

df_okexo2.to_csv(cwd+"/Volume_Data/IGNORE/OKEX_OI_Final.csv",index=True,index_label="DATE")
df_okexo2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/OKEX_OI_Base.csv",index=True,index_label="DATE")


##################HOUBI_OI#############################################

value=data["data"]["dataMap"]["Huobi"]

#huobi_oi=[]
#huobi_oi.append([time_new,value])



df_value=pd.DataFrame(value,columns=["HOUBI OI"])

huobi_df=pd.concat([df_time,df_value],axis=1)

huobi_df.to_csv(cwd+"/Volume_Data/IGNORE/houbi_OI_Final.csv",index=False,header=True,index_label="DATE")

x3 = pd.read_csv(cwd+"/Volume_Data/IGNORE/BASE/houbi_OI_Base.csv",parse_dates=["DATE"],dayfirst=True,index_col="DATE")
x4 = pd.read_csv(cwd+"/Volume_Data/IGNORE/houbi_OI_Final.csv", parse_dates=["DATE"], dayfirst=True,index_col="DATE")


huobi_final_hope_2=pd.concat([x3,x4]).reset_index()#.drop_duplicates(keep="last")
df4=pd.DataFrame(huobi_final_hope_2)
df4 = df4.dropna()
df4=df4.set_index("DATE")
df_huobio2= df4[~df4.index.duplicated()]

df_huobio2.to_csv(cwd+"/Volume_Data/IGNORE/houbi_OI_Final.csv",index=True,index_label="DATE")
df_huobio2.to_csv(cwd+"/Volume_Data/IGNORE/BASE/houbi_OI_Base.csv",index=True,index_label="DATE")


##excel work###
new_2=pd.concat([df_bno2,df_cmeo2,df_okexo2,df_huobio2],axis=1)
new_2= new_2.resample('D').asfreq()
new_2= new_2.interpolate(method='linear', axis=0).ffill().bfill()

append_df_to_excel(cwd+"//Volume_Data//Volumes_Combined//Volumes_Open_Interest_Combined.xlsx",new_2,sheet_name="OPEN_INTEREST_DAILY",startrow=0)
